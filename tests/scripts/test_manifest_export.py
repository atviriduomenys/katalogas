import csv
import hashlib
import shutil
from pathlib import Path
from typing import Iterable
from typing import Iterator

import pytest
import openpyxl
from factory.django import FileField
from django.db.models import QuerySet

from vitrina.datasets.factories import DatasetStructureFactory
from vitrina.structure.models import Metadata
from vitrina.datasets.models import Dataset
from vitrina.datasets.models import DatasetStructure
# from vitirna.scripts.manifest.exporter
# from vitirna.scripts.manifest.importer


@pytest.mark.parametrize('sample,sep', [
    ('id;dataset;resource', ';'),
    ('id,dataset,resource', ','),
    ('id\tdataset\tresource', '\t'),
])
def test_read_csv_table_detect_sep(sample: str, sep: str):
    assert detect_csv_sep(sample) == sep


def detect_csv_sep(sample: str) -> str:
    for sep in (',', ';', '\t'):
        if sep in sample:
            return sep


def write_csv_table(path: Path, table: Iterable[list[str]]):
    with open(path, 'w', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row in table:
            writer.writerow(row)


def read_csv_table(path: Path) -> Iterator[list[str]]:
    with open(path, encoding='utf-8') as f:
        sample = f.read(20)
    sep = detect_csv_sep(sample)

    with open(path, encoding='utf-8') as f:
        if sep:
            reader = csv.reader(f, delimiter=sep, quotechar='|')
            for row in reader:
                yield [v.strip() for v in row]
        else:
            for row in f:
                yield row.strip()


def write_xlsx_table(path: Path, table: Iterable[list[str]]):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in table:
        ws.append(row)
    wb.save(path)


def read_xlsx_table(path: Path) -> Iterator[list[str]]:
    wb = openpyxl.load_workbook(path)
    for sheet in wb:
        buffer = []
        rows = sheet.iter_rows(max_col=20, values_only=True)
        header = next(rows, None)
        header = [str(v).strip() if v else '' for v in header]
        header = rstrip_list(header)
        yield header
        for row in rows:
            row = [str(v).strip() if v else '' for v in row]
            row = rstrip_list(row, limit=len(header))
            buffer.append(row)
            if any(row):
                yield from buffer
                buffer = []
            elif len(buffer) > 100:
                break


@pytest.mark.parametrize('lst, result', [
    ('1,2,3', '1,2,3'),
    ('1,2,3,', '1,2,3'),
    ('1,2,3,,', '1,2,3'),
    (',1,2,3,,', ',1,2,3'),
])
def test_rstrip_list(lst, result):
    assert ','.join(rstrip_list(lst.split(','))) == result


@pytest.mark.parametrize('lst, result', [
    ('1,2', '1,2'),
    ('1,2,,,,', '1,2,,'),
])
def test_rstrip_list_limit(lst, result):
    assert ','.join(rstrip_list(lst.split(','), limit=4)) == result


def rstrip_list(value, *, limit: int | None = None):
    size = len(value)
    pos = next(
        (i for i, v in enumerate(reversed(value)) if v),
        size - 1,
    )
    pos = size - pos
    if limit is not None and size > limit:
        pos = limit
    return value[:pos]


def read_table(format: str, path: Path):
    reader = FORMATS[format]['reader']
    yield from reader(path)


def write_table(format: str, path: Path, table: Iterable[list[str]]):
    writer = FORMATS[format]['writer']
    writer(path, table)


def get_format_from_path(path: Path):
    return path.suffix.strip('.')


FORMATS = {
    'csv': {
        'reader': read_csv_table,
        'writer': write_csv_table,
    },
    'xlsx': {
        'reader': read_xlsx_table,
        'writer': write_xlsx_table,
    },
}


@pytest.mark.parametrize('format', list(FORMATS))
def test_table_io(tmp_path: Path, format: str):
    table = [
        ['id', 'dataset', 'resource'],
        ['', 'datasets/gov/ivpk/adp', ''],
    ]

    path = (tmp_path / f'manifest.{format}')
    write_table(format, path, table)
    assert list(read_table(format, path)) == table


def test_table_to_dicts():
    table = [
        ['id', 'dataset', 'resource'],
        ['', 'datasets/gov/ivpk/adp', ''],
    ]
    assert list(iter_table_as_dicts(table)) == [
        {'id': '', 'dataset': 'datasets/gov/ivpk/adp', 'resource': ''},
    ]


def iter_table_as_dicts(table: Iterable[list[str]]):
    table = iter(table)
    header = next(table, None)
    if header is None:
        return
    for row in table:
        yield {k: v for k, v in zip(header, row)}


@pytest.mark.parametrize('table, name', [
    ([{'dataset': 'datasets/gov/ivpk/adp'}], 'datasets/gov/ivpk/adp'),
    ([{'dataset': ''}], None),
    ([{'id': ''}], None),
])
def test_get_dataset_name_from_tabular(table, name):
    assert get_dataset_name_from_table(table) == name


def get_dataset_name_from_table(table: Iterable[dict[str, str]]) -> str:
    for row in table:
        if 'dataset' not in row:
            return
        if row['dataset']:
            return row['dataset']


@pytest.mark.parametrize('dataset, org', [
    ('datasets/gov/ivpk/adp', 'ivpk'),
    ('', None),
    (None, None),
])
def test_get_org_name_from_dataset_name(dataset, org):
    assert get_org_name_from_dataset_name(dataset) == org


def get_org_name_from_dataset_name(dataset: str | None) -> str | None:
    if dataset and dataset.startswith('datasets/gov/'):
        parts = dataset.split('/')
        return parts[2]


@pytest.mark.django_db
def test_export(tmp_path: Path):
    structure = DatasetStructureFactory(
        dataset__organization__title="Org 1",
        dataset__title="Dataset 1",
        file__file=FileField(filename='manifest.csv', data=(
            b'id,dataset,resource\n'
            b',datasets/gov/ivpk/adp,\n'
        )),
    )
    datasets = query_datasets_with_metadata()
    (dataset, metadata), = datasets

    path = Path(structure.file.path)
    format = get_format_from_path(path)
    table = read_table(format, path)
    table = iter_table_as_dicts(table)
    dataset_name = get_dataset_name_from_table(table)
    assert dataset_name == 'datasets/gov/ivpk/adp'
    org_name = get_org_name_from_dataset_name(dataset_name)
    assert org_name == 'ivpk'


def query_datasets_with_metadata(datasets: QuerySet[Dataset] = None):
    if datasets is None:
        datasets = Dataset.objects.all()
    datasets = (
        datasets.
        select_related(
            'organization',
            'current_structure',
        ).
        prefetch_related(
            'metadata',
        )
    )
    for dataset in datasets:
        try:
            metadata = dataset.metadata.get()
        except Metadata.DoesNotExist:
            metadata = Metadata()
        yield dataset, metadata


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        while chunk := f.read(h.block_size):
            h.update(chunk)
    return h.hexdigest()


@pytest.mark.parametrize('path', [
    'foo/bar/baz.csv',
    'datasets/gov/ivpk/adp.csv',
])
def test_fix_manifest_path(tmp_path: Path, path: str):
    path = tmp_path / path
    path.parent.mkdir(0o755, parents=True)
    write_csv_table(path, [
        ['id', 'dataset', 'resource'],
        ['', 'datasets/gov/ivpk/adp', ''],
    ])
    path = fix_manifest_path(tmp_path, path, 'datasets/gov/ivpk/adp')
    assert path == tmp_path / 'datasets/gov/ivpk/adp.csv'
    assert strip_base_from_paths(tmp_path, tmp_path.glob('**/*.csv')) == [
        Path('datasets/gov/ivpk/adp.csv'),
    ]


def fix_manifest_path(base: Path, path: Path, dataset: str) -> Path:
    correct_path = base / f'{dataset}.csv'
    if path.resolve() != correct_path.resolve():
        correct_path.parent.mkdir(0o755, parents=True, exist_ok=True)
        shutil.move(path, correct_path)
    return correct_path


def strip_base_from_paths(base: Path, paths: Iterable[Path]) -> list[Path]:
    return [path.relative_to(base) for path in paths]


def test_mapping_tables():
    orgmap = Mapping(['regno', 'title'])
    orgmap.update(42, None)
    orgmap.update(42, 'ORG')
    orgmap.update(None, 'ORG')
    assert orgmap.data == [
        {'id': 42, 'name': 'ORG', 'regno': '', 'title': ''},
    ]


class Mapping:
    names: list[str]
    by_id: dict[str, int]
    by_name: dict[str, int]

    def __init__(self, names: list[str]):
        self.names = names

    def update(self, id, name, **kwargs) -> None:
        id = str(id) if id else ''
        name = name if name else ''
        item = {
            'id': id,
            'name': name,
            **kwargs,
        }
        self.data.append(item)
        ix = len(self.data)
        if id:
            self.by_id[id] = ix
        if name:
            self.by_name[name] = ix

